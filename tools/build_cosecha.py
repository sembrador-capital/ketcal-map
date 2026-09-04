#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Consolida las planillas de cosecha de Ketcal en `cosecha_data.json`, la fuente
de la pestana Cosecha del mapa.

Decisiones que importan:

* La unidad espacial es el CUARTEL, y la clave es el par (especie, numero).
  Las planillas numeran los cuarteles por especie —Limon 1..14, Naranja 1..12,
  Mandarina 1..4— y el mapa los llama LIM-C1, NAR-C1, MAN-C1. Los 30 cruzan
  1 a 1 y las superficies declaradas calzan con la geometria del KMZ dentro
  del 2 % (la diferencia es cabecera y caminos, que el poligono incluye y la
  plantacion no).

* Las hectareas para el rendimiento son las PLANTADAS del cuadro de plantacion,
  no las geometricas del KMZ. Es el denominador que usa la propia planilla en
  su columna `Kg / Ha`, y es el agronomicamente correcto. Se usa el mismo valor
  para todas las temporadas para que kg/ha sea comparable entre anios: las
  planillas de 2025 y 2026 traen la misma superficie redondeada distinto
  (5,91 vs 5,92) y eso solo mueve el tercer decimal.

* La semana sale de la columna `Semana` de la planilla, NO del calendario ISO
  de la fecha. Difieren en 2 filas de 687 en 2025, ambas domingos que el packing
  cuenta en la semana siguiente. La convencion operativa del campo gana; las
  discrepancias quedan en `issues`.

* El destino se reduce a TRES clases comparables entre temporadas
  —exportacion / mercado interno / desecho— porque 2025 no separa el camote y
  2026 si. El camote se guarda aparte como desglose del mercado interno, que es
  lo que agronomicamente es: fruta de menor calibre que se vende igual.

* `Cuartel Real` (solo en 2026) NO se usa como clave. En las 23 filas donde
  difiere de `Cuartel`, la superficie y el kg/ha de la propia fila siguen
  siempre a `Cuartel`, y `Cuartel Real` trae valores que no existen en limones
  (15, 20, 29). Las 23 estan en un unico bloque de tres dias: tiene la firma de
  un arrastre de formula. Quedan listadas en `issues` para que agronomia las
  revise; no se corrigen aca.

Uso:
    python tools/build_cosecha.py datos_fuente/*.xlsx [--geo geo_data.json]
                                  [--out cosecha_data.json]
"""

from __future__ import annotations

import argparse
import json
import re
import sys
import unicodedata
from collections import Counter, OrderedDict, defaultdict
from datetime import date, datetime, timezone
from pathlib import Path

import openpyxl

# Prefijo del id de cuartel en geo_data.json, por especie de la planilla.
ESPECIES = {
    "limon": ("LIM", "Limoneros", "Lim", "Limón", "Lemon"),
    "naranja": ("NAR", "Naranjos", "Nar", "Naranja", "Orange"),
    "mandarina": ("MAN", "Mandarinos", "Man", "Mandarina", "Mandarin"),
}

# Las tres clases comparables entre temporadas, en el orden en que se apilan.
DESTINOS = [
    ("exportacion",     "Exportación",    "Export",          "--dat-dest-exp"),
    ("mercado_interno", "Mercado interno", "Domestic market", "--dat-dest-mi"),
    ("desecho",         "Desecho",        "Waste",           "--dat-dest-des"),
]

# Como se lee cada valor de las planillas. 2025 usa la columna `Exp.-MI`
# y 2026 `Destino Fruta`; el vocabulario tambien cambia.
MAPA_DESTINO = {
    "exportacion": "exportacion",
    "exportación": "exportacion",
    "mercado interno": "mercado_interno",
    "minterno - camote": "mercado_interno",   # camote = mercado interno, se
    "minterno-camote": "mercado_interno",     # desglosa aparte
    "camote": "mercado_interno",
    "merma": "desecho",
    "desecho": "desecho",
    "botadero": "desecho",
}
ES_CAMOTE = ("minterno - camote", "minterno-camote", "camote")

MESES = ["enero", "febrero", "marzo", "abril", "mayo", "junio",
         "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre"]

# Receptores de la fruta, tal como los escribe la columna `Destino`. La planilla
# los tipea de varias formas —EL PARQUE / El Parque, ROSALES / Rosales— y sin
# canonizar salian como receptores distintos. El `id` es el que usa el mapa para
# buscar el logo en assets/exportadoras/<id>.png.
#
# `exportadora` distingue a quien recibe fruta de exportacion del botadero, que
# no es una exportadora sino el destino de la merma. Meterlos en la misma torta
# haria parecer que el desecho es un cliente.
RECEPTORES = {
    "rosales":              ("Rosales", "rosales", True),
    "gesex":                ("Gesex", "gesex", True),
    "propal":               ("Propal", "propal", True),
    "el parque":            ("El Parque", "el_parque", True),
    "westfalia":            ("Westfalia", "westfalia", True),
    "rio blanco":           ("Río Blanco", "rio_blanco", True),
    "inversion cordillera": ("Inversión Cordillera", "inversion_cordillera", True),
    "botadero":             ("Botadero", "botadero", False),
}


def receptor(txt):
    """(id, nombre, es_exportadora) canonico de un receptor."""
    k = sinacento(txt)
    if k in RECEPTORES:
        nombre, rid, expo = RECEPTORES[k]
        return rid, nombre, expo
    if not k:
        return None, None, False
    # Un receptor nuevo no se descarta: entra con su nombre tal cual y se
    # asume exportadora, que es lo que casi siempre sera.
    return slug_txt(k), norm(txt).title(), True


def slug_txt(s):
    return re.sub(r"[^a-z0-9]+", "_", sinacento(s)).strip("_")


# ----------------------------------------------------------------- utilidades

def norm(s):
    if s is None:
        return ""
    return re.sub(r"\s+", " ", str(s)).strip()


def sinacento(s):
    s = unicodedata.normalize("NFKD", norm(s).lower())
    return "".join(c for c in s if not unicodedata.combining(c))


def num(v):
    """Float tolerante: coma decimal, celdas de texto, '-' y vacios."""
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    t = norm(v).replace(".", "").replace(",", ".") if re.fullmatch(
        r"-?\d{1,3}(\.\d{3})+(,\d+)?", norm(v)) else norm(v).replace(",", ".")
    if not t or t == "-":
        return None
    m = re.fullmatch(r"-?\d+(?:\.\d+)?", t)
    return float(m.group(0)) if m else None


def entero(v):
    n = num(v)
    return int(round(n)) if n is not None else None


def fecha_de(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})", norm(v))
    return date(*map(int, m.groups())) if m else None


def r2(v, d=2):
    return None if v is None else round(v, d)


# ------------------------------------------------------- lectura de planillas

def buscar_hoja_cosecha(wb):
    """Devuelve (hoja, fila_encabezado, encabezados) de la hoja de eventos.

    No se busca por nombre: 2025 la llama "Cosecha Limones campo" y 2026
    "BD Ketcal". Se busca por forma —una fila que tenga a la vez `Cuartel` y
    `Kg Cosecha`— para que una temporada nueva con otro nombre entre sola.
    """
    for ws in wb.worksheets:
        for i, fila in enumerate(ws.iter_rows(min_row=1, max_row=6,
                                              values_only=True)):
            celdas = [sinacento(c) for c in fila]
            if "cuartel" in celdas and "kg cosecha" in celdas:
                return ws.title, i, [norm(c) if c is not None else None
                                     for c in fila]
    return None, None, None


def buscar_cuadro_plantacion(wb):
    """Cuadro de plantacion: especie, numero, variedades, centro de costo,
    anio y superficie plantada. Es el maestro de hectareas."""
    for ws in wb.worksheets:
        for i, fila in enumerate(ws.iter_rows(min_row=1, max_row=10,
                                              values_only=True)):
            celdas = [sinacento(c) for c in fila]
            if "especie" in celdas and any(
                    c.startswith("superficie") for c in celdas if c):
                idx = {}
                for j, c in enumerate(celdas):
                    if not c:
                        continue
                    if c == "especie":
                        idx["especie"] = j
                    elif "cuartel" in c:
                        idx["numero"] = j
                    elif "variedad" in c:
                        idx["variedades"] = j
                    elif "centro de costo" in c:
                        idx["centro_costo"] = j
                    elif "plantacion" in c:
                        idx["anio"] = j
                    elif c.startswith("superficie"):
                        idx["ha"] = j
                if "numero" in idx and "ha" in idx:
                    return ws, i, idx
    return None, None, None


def filas_de(ws, fila_hdr, hdr):
    out = []
    for i, r in enumerate(ws.iter_rows(values_only=True)):
        if i <= fila_hdr:
            continue
        if all(c is None or norm(c) == "" for c in r):
            continue
        out.append((i + 1, dict(zip(hdr, r))))     # fila 1-based, como Excel
    return out


def col(hdr, *candidatos):
    """Primera columna cuyo nombre normalizado coincide con algun candidato."""
    mapa = {sinacento(h): h for h in hdr if h}
    for c in candidatos:
        if sinacento(c) in mapa:
            return mapa[sinacento(c)]
    return None


# --------------------------------------------------------------- construccion

def clave_cuartel(especie, numero):
    e = sinacento(especie)
    for k, meta in ESPECIES.items():
        if e.startswith(k[:5]):
            return "%s-C%d" % (meta[0], numero)
    return None


def acumulador():
    return {
        "kg": 0.0, "bins": 0.0, "camote": 0.0,
        "dest": {d[0]: 0.0 for d in DESTINOS},
        "fechas": set(), "semanas": set(),
        "variedades": Counter(),
        # receptor -> {kg, dest{...}, cuarteles, semanas}
        "receptores": defaultdict(lambda: {
            "kg": 0.0, "dest": {d[0]: 0.0 for d in DESTINOS},
            "cuarteles": set(), "semanas": set(), "nombre": None, "expo": True}),
    }


def receptores_de(a):
    """Receptores ordenados por kilos, con su reparto por destino."""
    out = []
    for rid, v in a["receptores"].items():
        out.append(OrderedDict([
            ("id", rid), ("nombre", v["nombre"]), ("exportadora", v["expo"]),
            ("kg", r2(v["kg"], 1)),
            ("dest", {k: r2(x, 1) for k, x in v["dest"].items()}),
            ("cuarteles", len(v["cuarteles"])),
            ("semanas", len(v["semanas"])),
        ]))
    out.sort(key=lambda x: -x["kg"])
    return out


def cerrar(a, ha=None):
    out = OrderedDict()
    out["kg"] = r2(a["kg"], 1)
    out["bins"] = r2(a["bins"], 1)
    out["dest"] = {k: r2(v, 1) for k, v in a["dest"].items()}
    if a["camote"]:
        out["camote"] = r2(a["camote"], 1)
    if a["kg"]:
        out["pct"] = {k: round(v / a["kg"], 4) for k, v in a["dest"].items()}
    else:
        out["pct"] = {k: 0.0 for k in a["dest"]}
    out["kg_bin"] = r2(a["kg"] / a["bins"], 1) if a["bins"] else None
    if ha:
        out["ha"] = r2(ha, 3)
        out["kg_ha"] = r2(a["kg"] / ha, 1)
    if a["fechas"]:
        out["desde"] = min(a["fechas"]).isoformat()
        out["hasta"] = max(a["fechas"]).isoformat()
        out["n_dias"] = len(a["fechas"])
    out["n_semanas"] = len(a["semanas"])
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx", nargs="+", help="planillas de cosecha (una por temporada)")
    ap.add_argument("--geo", default="geo_data.json")
    ap.add_argument("--out", default="cosecha_data.json")
    args = ap.parse_args()

    rutas = []
    for patron in args.xlsx:
        p = Path(patron)
        rutas.extend(sorted(p.parent.glob(p.name)) if "*" in p.name else [p])
    rutas = [p for p in rutas if p.exists() and not p.name.startswith("~$")]
    if not rutas:
        sys.exit("ERROR: no se encontro ninguna planilla")
    if not Path(args.geo).exists():
        sys.exit("ERROR: no existe %s" % args.geo)

    issues = []
    geo = json.loads(Path(args.geo).read_text(encoding="utf-8"))
    geo_cuarteles = {f["properties"]["id"]: f["properties"]
                     for f in geo["cuarteles"]["features"]}

    # ── Maestro de cuarteles: hectareas plantadas y ficha agronomica ────────
    plantacion = {}
    for ruta in rutas:
        wb = openpyxl.load_workbook(ruta, data_only=True, read_only=True)
        ws, fila_hdr, idx = buscar_cuadro_plantacion(wb)
        if ws is not None:
            for i, r in enumerate(ws.iter_rows(values_only=True)):
                if i <= fila_hdr:
                    continue
                esp = norm(r[idx["especie"]]) if "especie" in idx else ""
                n = entero(r[idx["numero"]])
                ha = num(r[idx["ha"]])
                if not esp or n is None or ha is None:
                    continue
                cid = clave_cuartel(esp, n)
                if not cid:
                    continue
                plantacion[cid] = {
                    "variedades": norm(r[idx["variedades"]]) if "variedades" in idx else None,
                    "centro_costo": norm(r[idx["centro_costo"]]) if "centro_costo" in idx else None,
                    "anio_plantacion": norm(r[idx["anio"]]) if "anio" in idx else None,
                    "ha_plantada": ha,
                }
        wb.close()

    # ── Temporadas ─────────────────────────────────────────────────────────
    temporadas = []
    vistos_cuartel = set()

    for ruta in rutas:
        wb = openpyxl.load_workbook(ruta, data_only=True, read_only=True)
        nombre_hoja, fila_hdr, hdr = buscar_hoja_cosecha(wb)
        if nombre_hoja is None:
            issues.append({"nivel": "warn", "archivo": ruta.name,
                           "msg": "sin hoja de eventos de cosecha; se omite"})
            wb.close()
            continue
        filas = filas_de(wb[nombre_hoja], fila_hdr, hdr)
        wb.close()

        c_temp = col(hdr, "Temporada")
        c_fec = col(hdr, "Fecha Cosecha")
        c_sem = col(hdr, "Semana")
        c_mes = col(hdr, "Mes")
        c_esp = col(hdr, "Especie")
        c_var = col(hdr, "Variedad", "Variedades")
        c_cua = col(hdr, "Cuartel")
        c_cc = col(hdr, "Centro de Costo")
        c_ha = col(hdr, "Superficie (ha)")
        c_bin = col(hdr, "Nº Bins", "N° Bins", "Bins")
        c_kg = col(hdr, "Kg Cosecha")
        c_dst = col(hdr, "Destino Fruta", "Exp.-MI")
        c_rec = col(hdr, "Destino")
        c_real = col(hdr, "Cuartel Real")

        if not (c_kg and c_cua and c_dst):
            issues.append({"nivel": "warn", "archivo": ruta.name,
                           "msg": "faltan columnas obligatorias; se omite"})
            continue

        temp = norm(filas[0][1].get(c_temp)) if c_temp and filas else ruta.stem
        tot = acumulador()
        por_cuartel = defaultdict(acumulador)
        por_semana = defaultdict(acumulador)
        por_cuartel_semana = defaultdict(lambda: defaultdict(acumulador))
        sin_cuartel = {"kg": 0.0, "bins": 0.0, "filas": 0}
        desalineados = []
        semana_no_iso = 0
        fuera_de_anio = []

        for nfila, r in filas:
            kg = num(r.get(c_kg)) or 0.0
            bins = num(r.get(c_bin)) or 0.0
            f = fecha_de(r.get(c_fec))
            sem = entero(r.get(c_sem))
            especie = norm(r.get(c_esp))
            numero = entero(r.get(c_cua))
            cid = clave_cuartel(especie, numero) if numero is not None else None

            crudo = sinacento(r.get(c_dst))
            destino = MAPA_DESTINO.get(crudo)
            if destino is None:
                issues.append({"nivel": "warn", "archivo": ruta.name,
                               "fila": nfila,
                               "msg": "destino no reconocido: %r" % norm(r.get(c_dst))})
                continue
            camote = kg if crudo in ES_CAMOTE else 0.0

            # La semana declarada manda; la ISO solo se usa para auditar.
            if f and sem is not None and sem != f.isocalendar()[1]:
                semana_no_iso += 1
            if f and temp.isdigit() and f.year != int(temp):
                fuera_de_anio.append({"fila": nfila, "fecha": f.isoformat(),
                                      "semana": sem, "kg": r2(kg, 1),
                                      "cuartel": cid})
                # Los kilos y la semana declarada se conservan; la fecha no.
                # Con ella dentro, el rango de la temporada 2025 arrancaba en
                # julio de 2023 por una sola fila mal tipeada.
                f = None
            if c_real:
                real = entero(r.get(c_real))
                if real is not None and numero is not None and real != numero:
                    desalineados.append({"fila": nfila,
                                         "fecha": f.isoformat() if f else None,
                                         "cuartel": cid, "cuartel_real": real,
                                         "kg": r2(kg, 1)})

            def sumar(a):
                a["kg"] += kg
                a["bins"] += bins
                a["dest"][destino] += kg
                a["camote"] += camote
                if f:
                    a["fechas"].add(f)
                if sem is not None:
                    a["semanas"].add(sem)
                v = norm(r.get(c_var))
                if v and not v.startswith("#"):
                    a["variedades"][v] += kg
                rid, rnombre, rexpo = receptor(r.get(c_rec))
                if rid:
                    ac = a["receptores"][rid]
                    ac["nombre"] = rnombre
                    ac["expo"] = rexpo
                    ac["kg"] += kg
                    ac["dest"][destino] += kg
                    if cid:
                        ac["cuarteles"].add(cid)
                    if sem is not None:
                        ac["semanas"].add(sem)

            sumar(tot)
            if sem is not None:
                sumar(por_semana[sem])
            if cid is None:
                sin_cuartel["kg"] += kg
                sin_cuartel["bins"] += bins
                sin_cuartel["filas"] += 1
                continue
            vistos_cuartel.add(cid)
            sumar(por_cuartel[cid])
            if sem is not None:
                sumar(por_cuartel_semana[cid][sem])

            # Si el cuadro de plantacion no trajo la superficie, se toma la de
            # la propia fila: sin denominador no hay rendimiento.
            if cid not in plantacion and c_ha:
                ha = num(r.get(c_ha))
                if ha:
                    plantacion[cid] = {"variedades": None, "centro_costo": norm(r.get(c_cc)),
                                       "anio_plantacion": None, "ha_plantada": ha}

        # ── Salida de la temporada ─────────────────────────────────────────
        ha_cosechada = sum(plantacion.get(c, {}).get("ha_plantada") or 0.0
                           for c in por_cuartel)
        t_out = OrderedDict()
        t_out["id"] = temp
        t_out["label"] = temp
        t_out["archivo"] = ruta.name
        t_out["hoja"] = nombre_hoja
        t_out["n_filas"] = len(filas)
        t_out["total"] = cerrar(tot, ha_cosechada)
        t_out["total"]["cuarteles"] = len(por_cuartel)
        t_out["receptores"] = receptores_de(tot)
        if sin_cuartel["filas"]:
            t_out["sin_cuartel"] = {"kg": r2(sin_cuartel["kg"], 1),
                                    "bins": r2(sin_cuartel["bins"], 1),
                                    "filas": sin_cuartel["filas"]}
        t_out["semanas"] = []
        for n in sorted(por_semana):
            a = por_semana[n]
            s = cerrar(a)
            s["n"] = n
            s["cuarteles"] = sum(1 for c in por_cuartel_semana
                                 if n in por_cuartel_semana[c])
            s["mes"] = MESES[min(a["fechas"]).month - 1] if a["fechas"] else None
            t_out["semanas"].append(s)
        t_out["semanas"].sort(key=lambda x: x["n"])

        t_out["cuarteles"] = OrderedDict()
        for cid in sorted(por_cuartel, key=lambda c: (c.split("-")[0],
                                                      int(c.split("-C")[1]))):
            a = por_cuartel[cid]
            ha = (plantacion.get(cid) or {}).get("ha_plantada")
            c_out = cerrar(a, ha)
            c_out["variedades"] = [v for v, _ in a["variedades"].most_common()]
            c_out["receptores"] = receptores_de(a)
            c_out["semanas"] = OrderedDict()
            for n in sorted(por_cuartel_semana[cid]):
                sa = por_cuartel_semana[cid][n]
                w = OrderedDict()
                w["kg"] = r2(sa["kg"], 1)
                w["bins"] = r2(sa["bins"], 1)
                w["dest"] = {k: r2(v, 1) for k, v in sa["dest"].items()}
                if sa["camote"]:
                    w["camote"] = r2(sa["camote"], 1)
                if sa["fechas"]:
                    w["desde"] = min(sa["fechas"]).isoformat()
                    w["hasta"] = max(sa["fechas"]).isoformat()
                c_out["semanas"][str(n)] = w
            t_out["cuarteles"][cid] = c_out

        temporadas.append(t_out)

        # ── Auditoria de la temporada ──────────────────────────────────────
        if sin_cuartel["filas"]:
            issues.append({
                "nivel": "warn", "temporada": temp,
                "msg": "%d fila(s) sin cuartel asignado (%s kg): entran en el "
                       "total del predio pero no pintan ningun poligono"
                       % (sin_cuartel["filas"], f"{sin_cuartel['kg']:,.0f}")})
        if semana_no_iso:
            issues.append({
                "nivel": "info", "temporada": temp,
                "msg": "%d fila(s) con semana declarada distinta de la ISO de su "
                       "fecha; manda la declarada, que es la convencion del packing"
                       % semana_no_iso})
        if fuera_de_anio:
            issues.append({
                "nivel": "warn", "temporada": temp,
                "msg": "%d fila(s) con fecha fuera del anio de la temporada"
                       % len(fuera_de_anio),
                "detalle": fuera_de_anio[:10]})
        if desalineados:
            issues.append({
                "nivel": "warn", "temporada": temp,
                "msg": "%d fila(s) donde `Cuartel Real` difiere de `Cuartel`. Se "
                       "usa `Cuartel`: la superficie y el kg/ha de la fila lo "
                       "siguen siempre. Revisar en la planilla."
                       % len(desalineados),
                "detalle": desalineados[:25]})

    # ── Catalogo de cuarteles ──────────────────────────────────────────────
    cuarteles = OrderedDict()
    for cid in sorted(geo_cuarteles, key=lambda c: (c.split("-")[0],
                                                    int(c.split("-C")[1]))):
        g = geo_cuarteles[cid]
        p = plantacion.get(cid, {})
        pre = cid.split("-")[0]
        meta = next((m for m in ESPECIES.values() if m[0] == pre), None)
        cuarteles[cid] = OrderedDict([
            ("id", cid),
            ("name", g.get("name")),
            ("etiqueta", g.get("etiqueta")),
            ("especie", meta[3] if meta else g.get("especie")),
            ("especie_en", meta[4] if meta else None),
            ("especie_corta", g.get("especie_corta")),
            ("numero", g.get("numero")),
            ("variedades", p.get("variedades") or " / ".join(g.get("variedades") or [])),
            ("centro_costo", p.get("centro_costo")),
            ("anio_plantacion", p.get("anio_plantacion")),
            ("ha_plantada", r2(p.get("ha_plantada"), 3)),
            ("ha_geo", r2(g.get("ha"), 2)),
            ("equipos", g.get("equipos")),
            ("cosechado", cid in vistos_cuartel),
        ])
        if p.get("ha_plantada") is None:
            issues.append({"nivel": "warn", "cuartel": cid,
                           "msg": "sin superficie plantada declarada; el mapa "
                                  "no puede calcular su rendimiento"})

    sin_cosecha = [c for c, v in cuarteles.items() if not v["cosechado"]]
    if sin_cosecha:
        issues.append({"nivel": "info",
                       "msg": "%d cuartel(es) sin ninguna fila de cosecha en "
                              "ninguna temporada: %s"
                              % (len(sin_cosecha), ", ".join(sin_cosecha))})

    temporadas.sort(key=lambda t: t["id"], reverse=True)

    payload = OrderedDict([
        ("generated_at", datetime.now(timezone.utc).isoformat(timespec="seconds")),
        ("source", [{"archivo": t["archivo"], "hoja": t["hoja"],
                     "temporada": t["id"], "filas": t["n_filas"]}
                    for t in temporadas]),
        ("nota_ha", "El rendimiento usa la superficie PLANTADA del cuadro de "
                    "plantacion, no la geometrica del KMZ."),
        ("destinos", [OrderedDict([("id", d[0]), ("es", d[1]), ("en", d[2]),
                                   ("token", d[3])]) for d in DESTINOS]),
        ("cuarteles", cuarteles),
        ("temporadas", temporadas),
        ("issues", issues),
    ])

    Path(args.out).write_text(
        json.dumps(payload, ensure_ascii=False, indent=1), encoding="utf-8")

    # ── Resumen en consola ─────────────────────────────────────────────────
    kb = Path(args.out).stat().st_size / 1024
    print("OK  %s  (%.0f KB)" % (args.out, kb))
    print("    %d cuarteles en el catalogo, %d con cosecha"
          % (len(cuarteles), sum(1 for v in cuarteles.values() if v["cosechado"])))
    for t in temporadas:
        tt = t["total"]
        print("    %s  %6.1f t  %5.0f bins  %5.1f t/ha  %2d cuarteles  "
              "%2d semanas (%s a %s)"
              % (t["id"], tt["kg"] / 1000, tt["bins"] or 0,
                 (tt.get("kg_ha") or 0) / 1000, tt["cuarteles"],
                 len(t["semanas"]), tt.get("desde"), tt.get("hasta")))
        print("        exportacion %4.1f %%   mercado interno %4.1f %%   desecho %4.1f %%"
              % tuple(tt["pct"][d[0]] * 100 for d in DESTINOS))
        expo = [r for r in t["receptores"] if r["exportadora"]]
        tot_e = sum(r["dest"]["exportacion"] for r in expo) or 1
        print("        exportadoras: " + " · ".join(
            "%s %.1f %%" % (r["nombre"], r["dest"]["exportacion"] / tot_e * 100)
            for r in sorted(expo, key=lambda x: -x["dest"]["exportacion"])
            if r["dest"]["exportacion"]))
    if issues:
        print("    ISSUES (%d):" % len(issues))
        for i in issues:
            print("      [%s] %s" % (i.get("nivel", "info"),
                                     i.get("msg", "")[:110]))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
