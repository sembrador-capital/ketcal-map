#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Consolida los dos registros de monitoreo de Ketcal en `pye_data.json`, la
fuente de la pestana "Control de PyE" (plagas y enfermedades).

Son dos monitoreos distintos y NO se mezclan en un solo numero:

* PLANTA — `Ketcal - Registro Monitoreo Planta.xlsx`, hoja `Historico`.
  Mensual, sobre el arbol, 30 cuarteles. La medida es la INCIDENCIA: que
  porcentaje de los arboles revisados tiene la plaga. Trae ademas el nivel de
  alerta del monitoreador, que resulto ser una funcion determinista de la
  incidencia (verificado sobre las 455 filas, sin un solo solapamiento):

      Normal  < 6 %      Bajo  6 - 10 %      Medio  10 - 20 %      Alto >= 20 %

  Esos cortes se guardan en el JSON y son los que pinta el mapa, para que la
  escala del mapa sea la MISMA que la del informe de campo.

* FRUTO — `Registro Monitoreo Fruto Ketcal 2026.xlsx`, hoja `BASE DE DATOS`.
  Control de calidad en cosecha, por semana y contratista. La medida es el
  porcentaje de FRUTOS con cada defecto sobre los evaluados. Cada muestra son
  50 frutos (verificado: los 175 grupos con porcentaje derivable dan 50 exacto),
  asi que el denominador de cualquier agregado es 50 x numero de muestras.

  Se usa `BASE DE DATOS` y no `BD Ketcal `: esta ultima es una copia parcial
  y vieja —2.113 filas contra 3.427, semanas 24-27 contra 24-33— y ademas
  tiene 130 filas con el mes 9 sobre fechas de junio y julio.

Decisiones que importan:

* La clave espacial es el cuartel del mapa (LIM-C#, NAR-C#, MAN-C#). Planta los
  nombra `L1-C8`, `L2-C1`, `Naranjos N1-C3`, `Tango M1-C2` —centro de costo mas
  numero— y Fruto los nombra por especie mas numero, como las planillas de
  cosecha. Los dos cruzan 1 a 1 y la particion por centro de costo calza con el
  cuadro de plantacion (Limones 1 = 7..11, Limones 2 = 1..6 y 12..14).

* En Fruto solo se guardan las filas con al menos un fruto afectado. Un defecto
  que no aparece en una muestra es un defecto que se busco y no se encontro: el
  denominador no sale de contar filas sino de las muestras, que se guardan
  aparte. Eso baja el archivo a un tercio sin perder un solo dato.

* Los defectos se clasifican en plaga / enfermedad / otro. La pestana es de
  plagas y enfermedades, pero el registro de fruta es de calidad y la mayoria de
  sus defectos son fisiologicos o de manejo. Meterlos todos en la misma bolsa
  haria parecer plaga a un golpe de sol. El russet queda como `otro` a
  proposito: en citricos puede ser fisiologico o de acaro y el registro no lo
  distingue.

Uso:
    python tools/build_pye.py [--planta RUTA] [--fruto RUTA]
                              [--geo geo_data.json] [--out pye_data.json]
"""

from __future__ import annotations

import argparse
import json
import re
import sys
import unicodedata
from collections import Counter, OrderedDict, defaultdict
from datetime import date, datetime, timezone
from pathlib import Path

import openpyxl

MESES_ES = ["ene", "feb", "mar", "abr", "may", "jun",
            "jul", "ago", "sep", "oct", "nov", "dic"]
MESES_EN = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
            "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

# Prefijo del cuartel en geo_data.json por especie declarada.
PREFIJO = {"limonero": "LIM", "limon": "LIM",
           "naranjo": "NAR", "naranja": "NAR",
           "mandarino": "MAN", "mandarina": "MAN"}

# Centro de costo de Planta -> prefijo. `L1`/`L2` y `N1`/`N2` son particiones
# del mismo rango de numeros, no numeraciones propias.
CC_PREFIJO = {"l1": "LIM", "l2": "LIM", "n1": "NAR", "n2": "NAR", "m1": "MAN"}

# Niveles de alerta del monitoreo, del informe de campo. El corte superior es
# exclusivo. Se emiten al JSON: el mapa no inventa ninguno.
NIVELES = [
    ("normal", "Normal", "Normal", 0.0, 6.0),
    ("bajo",   "Bajo",   "Low",    6.0, 10.0),
    ("medio",  "Medio",  "Medium", 10.0, 20.0),
    ("alto",   "Alto",   "High",   20.0, None),
]

# Clase de cada grupo de plaga de Planta. Lo que no este listado queda `plaga`.
CLASE_PLANTA = {
    "pudriciones": "enfermedad",
    "fumagina": "enfermedad",
    "oidio": "enfermedad",
}

PLAGA_EN = {
    "aranita": "Red spider mite", "mosquita blanca": "Whitefly",
    "proeulias": "Proeulia", "pudriciones": "Rots", "trips": "Thrips",
    "acaro ancho": "Broad mite", "acaro yema": "Bud mite",
    "conchuela": "Soft scale", "chanchito blanco": "Mealybug",
    "fumagina": "Sooty mould", "pulgon": "Aphid", "escama": "Armoured scale",
    "capachito": "Weevil", "oidio": "Powdery mildew",
}

# Clase de cada defecto de Fruto. El resto queda `otro`.
CLASE_FRUTO = {
    "acaro de la yema leve": "plaga", "acaro de la yema descartable": "plaga",
    "acaro ancho": "plaga", "escama blanca": "plaga", "escama roja": "plaga",
    "chanchito blanco": "plaga", "dano eulia": "plaga", "trips": "plaga",
    "botritis": "enfermedad", "gomosis": "enfermedad", "fumagina": "enfermedad",
}
FRUTO_EN = {
    "russet severo": "Severe russet", "russet leve": "Light russet",
    "golpe sol": "Sunburn", "sin roseta": "No calyx", "dano roseta": "Calyx damage",
    "acaro de la yema leve": "Bud mite (light)",
    "acaro de la yema descartable": "Bud mite (cull)",
    "acaro ancho": "Broad mite", "redondo": "Round", "rugoso": "Rough",
    "pedunculo largo": "Long stem", "herida abierta": "Open wound",
    "herida cicatrizada": "Healed wound", "pistilo": "Pistil",
    "bajo calibre": "Undersized", "sobre calibre": "Oversized",
    "protuberancia": "Protuberance", "deshidratado": "Dehydrated",
    "acostillado": "Ribbed", "oleocelosis": "Oleocellosis",
    "fumagina": "Sooty mould", "cuello botella": "Bottleneck",
    "dano tijera": "Clipper cut", "aleta pedunculo": "Stem wing",
    "desgarro": "Tear", "visos verdes": "Green tinge",
    "fruta del suelo": "Ground fruit", "escama blanca": "White scale",
    "escama roja": "Red scale", "chanchito blanco": "Mealybug",
    "dano eulia": "Eulia damage", "ombligo abierto": "Open navel",
    "sedimentos": "Sediment", "gomosis": "Gummosis", "trips": "Thrips",
    "pitting": "Pitting", "water spot": "Water spot", "botritis": "Botrytis",
}

# Variantes que la planilla escribe de dos maneras y que el slug no unifica
# porque difieren en letras, no en mayusculas ni acentos.
ALIAS_FRUTO = {"botrytis": "botritis"}

MERCADOS = {"exportacion": ("Exportación", "Export"),
            "mercado interno": ("Mercado interno", "Domestic market"),
            "basura": ("Basura", "Discard")}


# ----------------------------------------------------------------- utilidades

def norm(s):
    if s is None:
        return ""
    return re.sub(r"\s+", " ", str(s)).strip()


def sinacento(s):
    # El grado y el ordinal masculino se colapsan a "o" antes de normalizar:
    # las planillas escriben la misma columna como "N Frutos", "Nº Frutos" y
    # "N° Frutos", y NFKD solo descompone uno de los dos signos. Sin esto,
    # `col()` no encontraba la columna de frutos y todo el conteo daba cero.
    s = norm(s).lower().replace("º", "o").replace("°", "o")
    s = unicodedata.normalize("NFKD", s)
    return "".join(c for c in s if not unicodedata.combining(c))


def slug(s):
    return re.sub(r"[^a-z0-9]+", "_", sinacento(s)).strip("_")


def num(v):
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    t = norm(v).replace(",", ".")
    if not t or t == "-":
        return None
    m = re.fullmatch(r"-?\d+(?:\.\d+)?", t)
    return float(m.group(0)) if m else None


def entero(v):
    n = num(v)
    return int(round(n)) if n is not None else None


def fecha_de(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})", norm(v))
    return date(*map(int, m.groups())) if m else None


def r2(v, d=2):
    return None if v is None else round(v, d)


def hoja_filas(ruta, nombre, fila_hdr, off=0):
    wb = openpyxl.load_workbook(ruta, data_only=True, read_only=True)
    if nombre not in wb.sheetnames:
        # Tolerar el espacio final del nombre de hoja y las mayusculas.
        cand = [h for h in wb.sheetnames if sinacento(h) == sinacento(nombre)]
        if not cand:
            wb.close()
            return None, []
        nombre = cand[0]
    ws = wb[nombre]
    filas = list(ws.iter_rows(values_only=True))
    wb.close()
    hdr = [norm(c) or None for c in filas[fila_hdr][off:]]
    out = []
    for i, r in enumerate(filas[fila_hdr + 1:]):
        r = r[off:]
        if all(c is None or norm(c) == "" for c in r):
            continue
        out.append((fila_hdr + i + 2, dict(zip(hdr, r))))   # fila 1-based
    return hdr, out


def col(hdr, *cands):
    mapa = {sinacento(h): h for h in hdr if h}
    for c in cands:
        if sinacento(c) in mapa:
            return mapa[sinacento(c)]
    for c in cands:
        for k, v in mapa.items():
            if k.startswith(sinacento(c)[:8]):
                return v
    return None


def nivel_de(inc):
    """Nivel de alerta a partir de la incidencia, con los cortes del informe."""
    if inc is None:
        return None
    for clave, _, _, lo, hi in NIVELES:
        if inc >= lo and (hi is None or inc < hi):
            return clave
    return NIVELES[-1][0]


def etiqueta_mes(mes):
    y, m = mes.split("-")
    return "%s %s" % (MESES_ES[int(m) - 1], y), "%s %s" % (MESES_EN[int(m) - 1], y)


# ------------------------------------------------------------------- cuarteles

RX_PLANTA = re.compile(r"(?:^|\s)(L1|L2|N1|N2|M1)\s*-\s*C\s*(\d+)\s*$", re.I)


def cuartel_planta(txt):
    """`L2-C1`, `Naranjos N1-C3`, `Tango M1-C2` -> LIM-C1 / NAR-C3 / MAN-C2."""
    m = RX_PLANTA.search(norm(txt))
    if not m:
        return None
    pre = CC_PREFIJO.get(m.group(1).lower())
    return "%s-C%d" % (pre, int(m.group(2))) if pre else None


def cuartel_fruto(especie, numero):
    pre = PREFIJO.get(sinacento(especie))
    return "%s-C%d" % (pre, numero) if pre and numero is not None else None


# --------------------------------------------------------------------- planta

def leer_planta(ruta, geo, issues):
    hdr, filas = hoja_filas(ruta, "Historico", 0)
    if not filas:
        issues.append({"nivel": "warn", "fuente": "planta",
                       "msg": "no se encontro la hoja Historico"})
        return None

    c_mes = col(hdr, "mes")
    c_cua = col(hdr, "cuartel")
    c_cc = col(hdr, "cc")
    c_esp = col(hdr, "especie")
    c_var = col(hdr, "variedad")
    c_pl = col(hdr, "plaga_nombre")
    c_gr = col(hdr, "grupo_plaga")
    c_nm = col(hdr, "n_arboles_monitoreados")
    c_np = col(hdr, "n_arboles_presencia")
    c_inc = col(hdr, "incidencia_pct")
    c_ab = col(hdr, "nivel_abundancia")
    c_al = col(hdr, "nivel_alerta")

    registros = []
    plagas = OrderedDict()
    meses = Counter()
    sin_cuartel = 0
    alerta_discrepante = []

    for nfila, r in filas:
        cid = cuartel_planta(r.get(c_cua))
        if not cid or cid not in geo:
            sin_cuartel += 1
            issues.append({"nivel": "warn", "fuente": "planta", "fila": nfila,
                           "msg": "cuartel sin equivalencia en el mapa: %r"
                                  % norm(r.get(c_cua))})
            continue
        mes = norm(r.get(c_mes))[:7]
        if not re.fullmatch(r"\d{4}-\d{2}", mes):
            issues.append({"nivel": "warn", "fuente": "planta", "fila": nfila,
                           "msg": "mes ilegible: %r" % norm(r.get(c_mes))})
            continue

        grupo = norm(r.get(c_gr)) or norm(r.get(c_pl))
        gid = slug(grupo)
        if not gid:
            continue
        inc = num(r.get(c_inc))
        nivel = nivel_de(inc)
        declarado = slug(r.get(c_al)) or None
        # El nivel declarado y el derivado de la incidencia coinciden en las 455
        # filas. Si alguna vez dejan de coincidir hay que verlo, no taparlo.
        if declarado and nivel and declarado != nivel:
            alerta_discrepante.append({"fila": nfila, "cuartel": cid, "mes": mes,
                                       "plaga": grupo, "incidencia": inc,
                                       "declarado": norm(r.get(c_al)),
                                       "derivado": nivel})

        if gid not in plagas:
            plagas[gid] = {"id": gid, "es": grupo.title() if grupo.isupper() else grupo,
                           "en": PLAGA_EN.get(sinacento(grupo)) or grupo.title(),
                           "clase": CLASE_PLANTA.get(sinacento(grupo), "plaga"),
                           "n": 0, "cuarteles": set(), "meses": set(),
                           "especies": set(), "nombres": Counter()}
        p = plagas[gid]
        p["n"] += 1
        p["cuarteles"].add(cid)
        p["meses"].add(mes)
        p["especies"].add(norm(r.get(c_esp)))
        if norm(r.get(c_pl)):
            p["nombres"][norm(r.get(c_pl))] += 1
        meses[mes] += 1

        registros.append(OrderedDict([
            ("c", cid), ("m", mes), ("p", gid),
            ("n", entero(r.get(c_nm))), ("pr", entero(r.get(c_np))),
            ("i", r2(inc, 1)), ("ab", r2(num(r.get(c_ab)), 4)),
            ("a", nivel),
            ("esp", norm(r.get(c_pl)) or None),
            ("cc", norm(r.get(c_cc)) or None),
            ("var", norm(r.get(c_var)) or None),
        ]))

    if alerta_discrepante:
        issues.append({"nivel": "warn", "fuente": "planta",
                       "msg": "%d fila(s) donde el nivel de alerta declarado no "
                              "coincide con el que sale de la incidencia; manda "
                              "el declarado" % len(alerta_discrepante),
                       "detalle": alerta_discrepante[:12]})
    if sin_cuartel:
        issues.append({"nivel": "warn", "fuente": "planta",
                       "msg": "%d fila(s) descartadas por cuartel desconocido"
                              % sin_cuartel})

    cat = []
    for gid, p in sorted(plagas.items(), key=lambda kv: -kv[1]["n"]):
        cat.append(OrderedDict([
            ("id", gid), ("es", p["es"]), ("en", p["en"]), ("clase", p["clase"]),
            ("n", p["n"]),
            ("cuarteles", len(p["cuarteles"])),
            ("meses", sorted(p["meses"])),
            ("especies", sorted(x for x in p["especies"] if x)),
            ("nombres", [k for k, _ in p["nombres"].most_common()]),
        ]))

    lista_meses = []
    for m in sorted(meses):
        es, en = etiqueta_mes(m)
        cu = {r["c"] for r in registros if r["m"] == m}
        pl = {r["p"] for r in registros if r["m"] == m}
        lista_meses.append(OrderedDict([
            ("id", m), ("es", es), ("en", en), ("n", meses[m]),
            ("cuarteles", len(cu)), ("plagas", len(pl))]))

    return OrderedDict([
        ("archivo", Path(ruta).name),
        ("hoja", "Historico"),
        ("n_registros", len(registros)),
        ("niveles", [OrderedDict([("id", a), ("es", b), ("en", c),
                                  ("min", lo), ("max", hi)])
                     for a, b, c, lo, hi in NIVELES]),
        ("meses", lista_meses),
        ("plagas", cat),
        ("registros", registros),
    ])


# ---------------------------------------------------------------------- fruto

def leer_fruto(ruta, geo, issues):
    hdr, filas = hoja_filas(ruta, "BASE DE DATOS", 0)
    if not filas:
        issues.append({"nivel": "warn", "fuente": "fruto",
                       "msg": "no se encontro la hoja BASE DE DATOS"})
        return None

    c_fec = col(hdr, "FECHA")
    c_sem = col(hdr, "SEMANA")
    c_esp = col(hdr, "ESPECIE")
    c_cua = col(hdr, "CUARTEL")
    c_ct = col(hdr, "CONTRATISTA")
    c_mk = col(hdr, "MERCADO BINS")
    c_def = col(hdr, "DAÑOS DEFECTOS", "DANOS DEFECTOS")
    c_n = col(hdr, "Nº FRUTOS", "N FRUTOS")
    c_pct = col(hdr, "%")

    muestras = OrderedDict()      # clave -> indice
    lista_muestras = []
    registros = []
    defectos = OrderedDict()
    descartadas = 0
    frutos_huerfanos = 0.0
    tam_raros = Counter()

    for nfila, r in filas:
        cid = cuartel_fruto(r.get(c_esp), entero(r.get(c_cua)))
        sem = entero(r.get(c_sem))
        f = fecha_de(r.get(c_fec))
        if not cid or cid not in geo or sem is None or f is None:
            descartadas += 1
            frutos_huerfanos += num(r.get(c_n)) or 0.0
            continue

        mk = sinacento(r.get(c_mk))
        mk = mk if mk in MERCADOS else (mk or "exportacion")
        ct = norm(r.get(c_ct)) or None
        clave = (cid, f.isoformat(), sem, mk, ct)
        if clave not in muestras:
            muestras[clave] = len(lista_muestras)
            lista_muestras.append(OrderedDict([
                ("c", cid), ("f", f.isoformat()), ("s", sem),
                ("mk", mk), ("ct", ct), ("n", 0)]))
        mi = muestras[clave]

        # Tamano de la muestra, deducido de frutos / porcentaje. Es 50 en todos
        # los grupos donde se puede derivar; se comprueba y se registra si no.
        nf = num(r.get(c_n)) or 0.0
        pct = num(r.get(c_pct)) or 0.0
        if nf > 0 and pct > 0:
            tam = int(round(nf / pct))
            tam_raros[tam] += 1
            if not lista_muestras[mi]["n"]:
                lista_muestras[mi]["n"] = tam

        etiqueta = norm(r.get(c_def))
        if not etiqueta or etiqueta.lower() == "none":
            continue
        did = ALIAS_FRUTO.get(slug(etiqueta), slug(etiqueta))
        if did not in defectos:
            base = sinacento(etiqueta)
            defectos[did] = {"id": did,
                             # La planilla escribe el mismo defecto en tres
                             # capitalizaciones distintas; gana la mas frecuente.
                             "formas": Counter(),
                             "en": FRUTO_EN.get(base) or etiqueta.capitalize(),
                             "clase": CLASE_FRUTO.get(base, "otro"),
                             "n": 0, "frutos": 0.0, "cuarteles": set()}
        d = defectos[did]
        d["formas"][etiqueta] += 1
        d["n"] += 1
        d["frutos"] += nf
        if nf > 0:
            d["cuarteles"].add(cid)
            # `nf` no siempre es entero: 768 filas traen medios frutos porque
            # promedian dos submuestras de 50. Truncar a entero perdia 64 frutos
            # repartidos por todos los defectos.
            registros.append(OrderedDict([("mi", mi), ("d", did), ("n", r2(nf, 1))]))

    # Muestras sin tamano derivable: se les pone el estandar y se avisa.
    estandar = tam_raros.most_common(1)[0][0] if tam_raros else 50
    sin_tam = [i for i, m in enumerate(lista_muestras) if not m["n"]]
    for i in sin_tam:
        lista_muestras[i]["n"] = estandar
    if sin_tam:
        issues.append({"nivel": "info", "fuente": "fruto",
                       "msg": "%d muestra(s) sin defectos con porcentaje: se les "
                              "asigna el tamano estandar de %d frutos"
                              % (len(sin_tam), estandar)})
    if len(tam_raros) > 1:
        issues.append({"nivel": "warn", "fuente": "fruto",
                       "msg": "la muestra no siempre tiene el mismo tamano: %s"
                              % dict(tam_raros)})
    if descartadas:
        issues.append({"nivel": "warn", "fuente": "fruto",
                       "msg": "%d fila(s) descartadas por cuartel, semana o fecha "
                              "ilegibles; %s fruto(s) afectados quedan fuera del "
                              "mapa" % (descartadas, r2(frutos_huerfanos, 1))})

    cat = []
    for did, d in sorted(defectos.items(), key=lambda kv: -kv[1]["frutos"]):
        es = d["formas"].most_common(1)[0][0]
        cat.append(OrderedDict([
            ("id", did), ("es", es), ("en", d["en"]), ("clase", d["clase"]),
            ("n", d["n"]), ("frutos", r2(d["frutos"], 1)),
            ("cuarteles", len(d["cuarteles"])),
            ("formas", sorted(d["formas"])) if len(d["formas"]) > 1 else ("formas", []),
        ]))

    semanas = OrderedDict()
    for m in lista_muestras:
        s = semanas.setdefault(m["s"], {"n": 0, "frutos": 0, "cuarteles": set(),
                                        "desde": m["f"], "hasta": m["f"]})
        s["n"] += 1
        s["frutos"] += m["n"]
        s["cuarteles"].add(m["c"])
        s["desde"] = min(s["desde"], m["f"])
        s["hasta"] = max(s["hasta"], m["f"])
    lista_sem = [OrderedDict([("n", k), ("muestras", v["n"]), ("frutos", v["frutos"]),
                              ("cuarteles", len(v["cuarteles"])),
                              ("desde", v["desde"]), ("hasta", v["hasta"])])
                 for k, v in sorted(semanas.items())]

    return OrderedDict([
        ("archivo", Path(ruta).name),
        ("hoja", "BASE DE DATOS"),
        ("n_filas", len(filas)),
        ("n_registros", len(registros)),
        ("mercados", [OrderedDict([("id", k), ("es", v[0]), ("en", v[1])])
                      for k, v in MERCADOS.items()]),
        ("semanas", lista_sem),
        ("defectos", cat),
        ("muestras", lista_muestras),
        ("registros", registros),
    ])


# ----------------------------------------------------------------------- main

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--planta", default="datos_fuente/Ketcal - Registro Monitoreo Planta.xlsx")
    ap.add_argument("--fruto", default="datos_fuente/Registro Monitoreo Fruto Ketcal 2026.xlsx")
    ap.add_argument("--geo", default="geo_data.json")
    ap.add_argument("--out", default="pye_data.json")
    args = ap.parse_args()

    if not Path(args.geo).exists():
        sys.exit("ERROR: no existe %s" % args.geo)
    geo = json.loads(Path(args.geo).read_text(encoding="utf-8"))
    geo_c = {f["properties"]["id"]: f["properties"]
             for f in geo["cuarteles"]["features"]}

    issues = []
    planta = leer_planta(args.planta, geo_c, issues) if Path(args.planta).exists() else None
    fruto = leer_fruto(args.fruto, geo_c, issues) if Path(args.fruto).exists() else None
    if planta is None and fruto is None:
        sys.exit("ERROR: no se pudo leer ninguna de las dos planillas")

    vistos = set()
    if planta:
        vistos |= {r["c"] for r in planta["registros"]}
    if fruto:
        vistos |= {m["c"] for m in fruto["muestras"]}

    cuarteles = OrderedDict()
    for cid in sorted(geo_c, key=lambda c: (c.split("-")[0], int(c.split("-C")[1]))):
        g = geo_c[cid]
        cuarteles[cid] = OrderedDict([
            ("id", cid), ("name", g.get("name")), ("etiqueta", g.get("etiqueta")),
            ("especie_corta", g.get("especie_corta")),
            ("variedades", " / ".join(g.get("variedades") or [])),
            ("ha", r2(g.get("ha"), 2)),
            ("planta", bool(planta) and any(r["c"] == cid for r in planta["registros"])),
            ("fruto", bool(fruto) and any(m["c"] == cid for m in fruto["muestras"])),
        ])

    sin = [c for c, v in cuarteles.items() if not v["planta"] and not v["fruto"]]
    if sin:
        issues.append({"nivel": "info",
                       "msg": "%d cuartel(es) sin ningun monitoreo: %s"
                              % (len(sin), ", ".join(sin))})

    payload = OrderedDict([
        ("generated_at", datetime.now(timezone.utc).isoformat(timespec="seconds")),
        ("source", [x for x in [
            planta and {"archivo": planta["archivo"], "hoja": planta["hoja"],
                        "fuente": "planta", "filas": planta["n_registros"]},
            fruto and {"archivo": fruto["archivo"], "hoja": fruto["hoja"],
                       "fuente": "fruto", "filas": fruto["n_filas"]}] if x]),
        ("clases", [
            OrderedDict([("id", "plaga"), ("es", "Plaga"), ("en", "Pest")]),
            OrderedDict([("id", "enfermedad"), ("es", "Enfermedad"), ("en", "Disease")]),
            OrderedDict([("id", "otro"), ("es", "Otro defecto"), ("en", "Other defect")]),
        ]),
        ("cuarteles", cuarteles),
        ("planta", planta),
        ("fruto", fruto),
        ("issues", issues),
    ])

    Path(args.out).write_text(
        json.dumps(payload, ensure_ascii=False, indent=1), encoding="utf-8")

    kb = Path(args.out).stat().st_size / 1024
    print("OK  %s  (%.0f KB)" % (args.out, kb))
    print("    %d cuarteles, %d con algun monitoreo"
          % (len(cuarteles), len(vistos)))
    if planta:
        print("    PLANTA  %d registros · %d cuarteles · %d meses (%s a %s) · %d plagas"
              % (planta["n_registros"],
                 len({r["c"] for r in planta["registros"]}),
                 len(planta["meses"]), planta["meses"][0]["id"],
                 planta["meses"][-1]["id"], len(planta["plagas"])))
        top = sorted(planta["plagas"], key=lambda p: -p["n"])[:6]
        for p in top:
            alto = sum(1 for r in planta["registros"]
                       if r["p"] == p["id"] and r["a"] == "alto")
            print("       %-20s %3d registros · %2d cuarteles · %2d en alerta alta"
                  % (p["es"][:20], p["n"], p["cuarteles"], alto))
    if fruto:
        frutos = sum(m["n"] for m in fruto["muestras"])
        print("    FRUTO   %d muestras · %d frutos · %d cuarteles · semanas %d-%d · %d defectos"
              % (len(fruto["muestras"]), frutos,
                 len({m["c"] for m in fruto["muestras"]}),
                 fruto["semanas"][0]["n"], fruto["semanas"][-1]["n"],
                 len(fruto["defectos"])))
        pye = [d for d in fruto["defectos"] if d["clase"] != "otro"]
        print("       %d defectos de plaga o enfermedad, %d frutos afectados"
              % (len(pye), sum(d["frutos"] for d in pye)))
    if issues:
        print("    ISSUES (%d):" % len(issues))
        for i in issues:
            print("      [%s] %s" % (i.get("nivel", "info"), i.get("msg", "")[:112]))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
